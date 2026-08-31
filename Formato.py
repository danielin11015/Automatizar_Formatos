# =====================================================================
# SISTEMA DE AUTOMATIZACIÓN DE REPORTES (BACKEND)
# Desarrollador: Erick Daniel Azamar Silva
# Institución: ESIME Culhuacán
# Descripción: API RESTful construida con Flask que recibe datos del 
# frontend, manipula Excel de forma invisible (COM) y genera PDFs.
# =====================================================================

from flask import Flask, render_template, request, jsonify
from flask_cors import CORS          # Habilito CORS para evitar bloqueos del navegador
import openpyxl                      # Librería para inyección súper rápida de datos en celdas
from PIL import Image                # Pillow: Motor de procesamiento matemático de imágenes
import os
import pythoncom                     # Requerido para abrir hilos de Windows COM en servidores web
import win32com.client               # Interfaz para manipular el proceso nativo de Microsoft Excel
from datetime import datetime
import gc                            # Garbage Collector: Para limpiar la RAM y matar procesos zombie
import base64                        # Para decodificar las imágenes que manda el frontend
import io

# Desactivo el límite de seguridad de tamaño de imagen. 
# Si el analista toma un recorte de 2 monitores, Pillow podría bloquearlo por error.
Image.MAX_IMAGE_PIXELS = None

app = Flask(__name__)
CORS(app)

# =====================================================================
# MOTOR MATEMÁTICO DE ACOMODO FOTOGRÁFICO (ASPECT RATIO)
# =====================================================================
def crear_mosaico(lista_imagenes_pil, target_aspect, nombre_salida, build_w=3200, tipo="normal"):
    if not lista_imagenes_pil: return None
    
    # Calculo la altura exacta del lienzo para que empate milimétricamente con Excel
    build_h = int(build_w / target_aspect)
    final_canvas = Image.new('RGB', (build_w, build_h), color='white')
    
    # --- MODO 1: POSICIONAL (Para Formato Móvil 3B) ---
    if tipo == "posicional":
        h_rojo = int(build_h * 0.38)     # 38% del alto (Visión en vivo)
        h_verde = int(build_h * 0.12)    # 12% del alto (Tiras de evidencia)
        h_medio = build_h - h_rojo - h_verde 
        w_azul = int(build_w * 0.45)     # 45% del ancho (Línea de tiempo)
        w_amarillo = build_w - w_azul    

        img_roja = lista_imagenes_pil[0] if len(lista_imagenes_pil) > 0 else None
        img_azul = lista_imagenes_pil[1] if len(lista_imagenes_pil) > 1 else None
        
        restantes = lista_imagenes_pil[2:] if len(lista_imagenes_pil) > 2 else []
        imgs_verdes, imgs_amarillas = [], []
        
        # Filtro Inteligente: Si el ancho es 4.5x más grande que el alto, es una tira larga.
        for img in restantes:
            if img.width > img.height * 4.5: imgs_verdes.append(img)
            else: imgs_amarillas.append(img) 

        # Inyecto imágenes con escalado BICUBIC para evitar que se pixelen
        if img_roja: final_canvas.paste(img_roja.convert("RGB").resize((build_w, h_rojo), Image.Resampling.BICUBIC), (0, 0))
        if img_azul: final_canvas.paste(img_azul.convert("RGB").resize((w_azul, h_medio), Image.Resampling.BICUBIC), (0, h_rojo))
        
        if imgs_amarillas:
            num_amarillas = len(imgs_amarillas)
            w_each = int(w_amarillo / num_amarillas)
            for i, img in enumerate(imgs_amarillas):
                w_current = w_each if i < num_amarillas - 1 else (w_amarillo - (w_each * i))
                pos_x = w_azul + (i * w_each)
                final_canvas.paste(img.convert("RGB").resize((w_current, h_medio), Image.Resampling.BICUBIC), (pos_x, h_rojo))
        
        if imgs_verdes:
            num_verdes = len(imgs_verdes)
            w_each_v = int(build_w / num_verdes)
            for i, img in enumerate(imgs_verdes):
                w_current_v = w_each_v if i < num_verdes - 1 else (build_w - (w_each_v * i))
                pos_x = i * w_each_v
                final_canvas.paste(img.convert("RGB").resize((w_current_v, h_verde), Image.Resampling.BICUBIC), (pos_x, h_rojo + h_medio))

    # --- MODO 2: DINÁMICO (Para Formato Torres Alcaldía) ---
    elif tipo == "torres_dinamico":
        if len(lista_imagenes_pil) > 0:
            img1 = lista_imagenes_pil[0]
            restantes = lista_imagenes_pil[2:] if len(lista_imagenes_pil) > 2 else []
            
            # Escenario A: La primera foto es panorámica (muy alargada)
            if img1.width > img1.height * 2.0:
                h_top = int(build_h * 0.5)      # Top toma la mitad superior
                h_bottom = build_h - h_top      
                w_left = int(build_w * 0.4)     # Cuadro pequeño izquierdo
                w_right = build_w - w_left      # Cuadro de grabación derecho
                
                final_canvas.paste(img1.convert("RGB").resize((build_w, h_top), Image.Resampling.BICUBIC), (0, 0))
                if len(lista_imagenes_pil) > 1:
                    img2 = lista_imagenes_pil[1]
                    final_canvas.paste(img2.convert("RGB").resize((w_right, h_bottom), Image.Resampling.BICUBIC), (w_left, h_top))
                
                if restantes:
                    h_each = int(h_bottom / len(restantes))
                    for i, img in enumerate(restantes):
                        h_c = h_each if i < len(restantes) - 1 else (h_bottom - h_each * i)
                        final_canvas.paste(img.convert("RGB").resize((w_left, h_c), Image.Resampling.BICUBIC), (0, h_top + (i * h_each)))
            
            # Escenario B: Las fotos son proporcionales (cuadradas/rectangulares normales)
            else:
                h_top = int(build_h * 0.5)      
                h_bottom = build_h - h_top      
                w_half = int(build_w * 0.5)     # Partimos a la mitad perfecta
                
                final_canvas.paste(img1.convert("RGB").resize((w_half, h_top), Image.Resampling.BICUBIC), (0, 0))
                if len(lista_imagenes_pil) > 1:
                    img2 = lista_imagenes_pil[1]
                    final_canvas.paste(img2.convert("RGB").resize((build_w - w_half, h_top), Image.Resampling.BICUBIC), (w_half, 0))
                
                if restantes:
                    w_each = int(build_w / len(restantes))
                    for i, img in enumerate(restantes):
                        w_c = w_each if i < len(restantes) - 1 else (build_w - w_each * i)
                        final_canvas.paste(img.convert("RGB").resize((w_c, h_bottom), Image.Resampling.BICUBIC), (i * w_each, h_top))

    # --- MODO 3: APILADO VERTICAL (Para celdas delgadas como IMEI) ---
    elif tipo == "apilado_vertical":
        num_imgs = len(lista_imagenes_pil)
        if num_imgs > 0:
            h_each = int(build_h / num_imgs)
            for i, img in enumerate(lista_imagenes_pil):
                h_current = h_each if i < num_imgs - 1 else (build_h - (h_each * i))
                final_canvas.paste(img.convert("RGB").resize((build_w, h_current), Image.Resampling.BICUBIC), (0, i * h_each))

    # --- MODO 4: MOSAICO TIPO PINTEREST (Página 2 / Vehículos) ---
    else:
        # Algoritmo matemático para distribuir fotos en filas (Ej: 3, 2, 2) y que cuadren perfecto
        imgs = [img.convert("RGB") for img in lista_imagenes_pil]
        margin = 15; num_imgs = len(imgs)
        best_rows, best_diff, best_distribution, best_total_w, best_total_h = 1, float('inf'), [num_imgs], 2000.0, 0.0

        for num_rows in range(1, num_imgs + 1):
            base_count = num_imgs // num_rows; extra = num_imgs % num_rows
            distribution = [(base_count + 1 if i < extra else base_count) for i in range(num_rows) if (base_count + 1 if i < extra else base_count) > 0]
            test_h, idx = 0.0, 0
            for count in distribution:
                row_imgs = imgs[idx : idx + count]
                idx += count
                total_aspect = sum((im.width / im.height) for im in row_imgs)
                if total_aspect == 0: total_aspect = 1
                test_h += (2000.0 - (count - 1) * margin) / total_aspect
            
            test_h += (len(distribution) - 1) * margin
            current_aspect = 2000.0 / test_h
            if abs(target_aspect - current_aspect) < best_diff:
                best_diff, best_distribution, best_total_h = abs(target_aspect - current_aspect), distribution, test_h

        cluster_canvas = Image.new('RGB', (2000, int(best_total_h)), color='white')
        current_y, idx = 0, 0
        for count in best_distribution:
            row_imgs = imgs[idx : idx + count]
            idx += count
            total_aspect = sum((im.width / im.height) for im in row_imgs)
            if total_aspect == 0: total_aspect = 1
            row_h = int((2000 - (count - 1) * margin) / total_aspect)
            current_x = 0
            for im in row_imgs:
                row_w = int(row_h * (im.width / im.height))
                cluster_canvas.paste(im.resize((row_w, row_h), Image.Resampling.BICUBIC), (current_x, current_y))
                current_x += row_w + margin
            current_y += row_h + margin
        final_canvas = cluster_canvas.resize((build_w, build_h), Image.Resampling.BICUBIC)

    final_canvas.save(nombre_salida, format="JPEG", quality=100, subsampling=0)
    return nombre_salida

# Decodifica la enorme cadena de texto Base64 del HTML y la convierte a bytes para Pillow
def decodificar_imagenes(lista_base64):
    imagenes_pil = []
    for b64_str in lista_base64:
        header, encoded = b64_str.split(",", 1)
        image_data = base64.b64decode(encoded)
        image_pil = Image.open(io.BytesIO(image_data))
        imagenes_pil.append(image_pil)
    return imagenes_pil


# =====================================================================
# RUTAS DE LA API FLASK (CONEXIÓN CON EL NAVEGADOR)
# =====================================================================
@app.route('/')
def home():
    return render_template('index.html')


# --------- ENDPOINT: FORMATO MÓVIL ---------
@app.route('/api/procesar_movil', methods=['POST'])
def procesar_movil():
    data = request.json
    pythoncom.CoInitialize() # Abro el hilo para controlar Excel sin corromper el servidor
    excel = None
    try:
        wb = openpyxl.load_workbook("Formato Validación movil (1).xlsx")
        ws = wb.active

        # Llenado rápido de datos
        ws['F7'] = data.get('cliente', ''); ws['J8'] = data.get('unidad', '')
        ws['F8'] = data.get('region', ''); ws['B8'] = data.get('fecha', '')
        ws['B11'] = data.get('h_ini', ''); ws['C11'] = data.get('h_fin', '')
        ws['F10'] = data.get('ticket', ''); ws['F62'] = data.get('monitorista', '')
        ws['F63'] = data.get('tecnico', ''); ws['B65'] = data.get('obs1', ''); ws['B118'] = data.get('obs2', '')

        # Lógica de nombre: MATRICULA_FECHA (Con mes en texto)
        meses = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
        raw_date = data.get('fecha', '')
        dt_obj = datetime.now()
        if raw_date:
            try: dt_obj = datetime.strptime(raw_date, '%Y-%m-%d')
            except: pass
        
        fecha_str = f"{dt_obj.day:02d}-{meses[dt_obj.month]}-{dt_obj.year}"
        matricula = str(data.get('unidad', 'SIN_MATRICULA')).strip().replace(" ", "_")
        nombre_reporte = f"{matricula}_{fecha_str}".upper()

        ruta_excel = os.path.abspath(f"{nombre_reporte}.xlsx")
        ruta_pdf = os.path.abspath(f"{nombre_reporte}.pdf")
        wb.save(ruta_excel); wb.close()

        # Abro instancia nativa de Microsoft Excel oculta
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False; excel.DisplayAlerts = False
        wb_win32 = excel.Workbooks.Open(ruta_excel)
        ws_win32 = wb_win32.Sheets(1)
        rutas_temp = []

        def inyectar(lista_b64, rango, nombre, res, tipo="normal"):
            if not lista_b64: return
            rx = ws_win32.Range(rango)
            r_img = crear_mosaico(decodificar_imagenes(lista_b64), rx.Width / rx.Height, nombre, res, tipo)
            if r_img:
                ra = os.path.abspath(r_img)
                ws_win32.Shapes.AddPicture(ra, False, True, rx.Left, rx.Top, rx.Width, rx.Height)
                rutas_temp.append(ra)

        # Inyección de todas las áreas
        inyectar(data.get('img_dvr', []), "F15:K33", f"tmp_dvr_{matricula}.jpg", 3200, tipo="posicional")
        inyectar(data.get('img_imei', []), "F36:K46", f"tmp_imei_{matricula}.jpg", 3200, tipo="apilado_vertical")
        inyectar(data.get('img_vehiculo', []), "B48:K61", f"tmp_vehiculo_{matricula}.jpg", 3200, tipo="normal")
        inyectar(data.get('img_pag2', []), "B78:K116", f"tmp_pag2_{matricula}.jpg", 3200, tipo="normal")

        wb_win32.Save()
        wb_win32.ExportAsFixedFormat(0, ruta_pdf)
        wb_win32.Close(False)

        # Convierto los archivos finales a Base64 para mandarlos en un solo JSON
        with open(ruta_pdf, "rb") as f_pdf: pdf_b64 = base64.b64encode(f_pdf.read()).decode('utf-8')
        with open(ruta_excel, "rb") as f_xls: xls_b64 = base64.b64encode(f_xls.read()).decode('utf-8')

        # Recolección de basura (Limpiar disco duro)
        os.remove(ruta_excel); os.remove(ruta_pdf)
        for r in rutas_temp:
            if os.path.exists(r): os.remove(r)
            
        return jsonify({"pdf": pdf_b64, "excel": xls_b64, "filename": nombre_reporte})

    except Exception as e: return jsonify({"error": str(e)}), 500
    finally:
        if excel: excel.Quit() # Destruyo el proceso EXCEL.EXE
        gc.collect()


# --------- ENDPOINT: FORMATO TORRES ---------
@app.route('/api/procesar_torres', methods=['POST'])
def procesar_torres():
    data = request.json
    pythoncom.CoInitialize()
    excel = None
    try:
        wb = openpyxl.load_workbook("Formato validacion torres.xlsx")
        ws = wb.active

        ws['A8'] = data.get('fecha', ''); ws['A11'] = data.get('h_ini', ''); ws['B11'] = data.get('h_fin', '')
        ws['G8'] = data.get('calle', ''); ws['G9'] = data.get('colonia', ''); ws['G10'] = data.get('id_torre', '')
        ws['G11'] = data.get('ticket', ''); ws['A45'] = data.get('tecnico', ''); ws['F45'] = data.get('monitorista', '')
        ws['A48'] = data.get('obs_generales', '')

        # Algoritmo de mapeo de Checklist (Lee las filas y pone la X donde debe)
        for item in data.get('checklist', []):
            nombre_buscar = item['buscar']; val = item['val']; obs = item['obs']
            for f in range(15, 35):
                celda = ws[f'A{f}'].value
                if celda and nombre_buscar.lower() in str(celda).lower():
                    if val == "No": ws[f'D{f}'] = ""; ws[f'E{f}'] = "X"
                    elif val == "N/A": ws[f'D{f}'] = "N/A"; ws[f'E{f}'] = ""
                    else: ws[f'D{f}'] = "X"; ws[f'E{f}'] = ""
                    if obs: ws[f'F{f}'] = obs
                    break

        meses = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
        raw_date = data.get('fecha', '')
        dt_obj = datetime.now()
        if raw_date:
            try: dt_obj = datetime.strptime(raw_date, '%Y-%m-%d')
            except: pass
        
        fecha_str = f"{dt_obj.day:02d}-{meses[dt_obj.month]}-{dt_obj.year}"
        matricula = str(data.get('id_torre', 'SIN_ID')).strip().replace(" ", "_")
        nombre_reporte = f"{matricula}_{fecha_str}".upper()

        ruta_excel = os.path.abspath(f"{nombre_reporte}.xlsx")
        ruta_pdf = os.path.abspath(f"{nombre_reporte}.pdf")
        wb.save(ruta_excel); wb.close()

        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False; excel.DisplayAlerts = False
        wb_win32 = excel.Workbooks.Open(ruta_excel)
        ws_win32 = wb_win32.Sheets(1)
        rutas_temp = []

        def inyectar(lista_b64, rango, nombre, res, tipo="normal"):
            if not lista_b64: return
            rx = ws_win32.Range(rango)
            r_img = crear_mosaico(decodificar_imagenes(lista_b64), rx.Width / rx.Height, nombre, res, tipo)
            if r_img:
                ra = os.path.abspath(r_img)
                ws_win32.Shapes.AddPicture(ra, False, True, rx.Left, rx.Top, rx.Width, rx.Height)
                rutas_temp.append(ra)

        # Llama a mi motor dinámico que escanea si la imagen es panorámica o cuadrada
        inyectar(data.get('img_evidencia', []), "A32:I42", f"tmp_torres_{matricula}.jpg", 1600, tipo="torres_dinamico")

        wb_win32.Save()
        wb_win32.ExportAsFixedFormat(0, ruta_pdf)
        wb_win32.Close(False)

        with open(ruta_pdf, "rb") as f_pdf: pdf_b64 = base64.b64encode(f_pdf.read()).decode('utf-8')
        with open(ruta_excel, "rb") as f_xls: xls_b64 = base64.b64encode(f_xls.read()).decode('utf-8')

        os.remove(ruta_excel); os.remove(ruta_pdf)
        for r in rutas_temp:
            if os.path.exists(r): os.remove(r)
            
        return jsonify({"pdf": pdf_b64, "excel": xls_b64, "filename": nombre_reporte})

    except Exception as e: return jsonify({"error": str(e)}), 500
    finally:
        if excel: excel.Quit()
        gc.collect()

if __name__ == '__main__':
    app.run(debug=True, port=5000)